#!/usr/bin/env python3
"""
Model evaluation script
"""

import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader
from torchvision import transforms
import pandas as pd
from PIL import Image
import os
import numpy as np
from sklearn.metrics import accuracy_score, precision_recall_fscore_support, cohen_kappa_score
from sklearn.metrics import confusion_matrix, classification_report
import matplotlib.pyplot as plt
# Configure matplotlib for English display
plt.rcParams['font.sans-serif'] = ['DejaVu Sans']  # Set English font
plt.rcParams['axes.unicode_minus'] = False  # Fix minus sign display
import seaborn as sns
import datetime
import json
import argparse

# 导入纯CoAtNet模型
from coatnet_training import create_pure_coatnet_model
from model_config import get_backbone_config


def find_image_file(images_dir, img_name):
    """Find image file, handle case issues, raise exception for invalid paths"""
    # Check if NaN or invalid value
    if pd.isna(img_name) or str(img_name).lower() == 'nan':
        raise ValueError(f"Invalid image path detected: NaN value")
    
    # First try direct path
    img_path = os.path.join(images_dir, str(img_name))
    if os.path.exists(img_path):
        return img_path
    
    # If direct path doesn't exist, try different extension variants
    base_name, ext = os.path.splitext(str(img_name))
    extensions = ['.jpg', '.JPG', '.jpeg', '.JPEG', '.png', '.PNG']
    
    for extension in extensions:
        potential_path = os.path.join(images_dir, base_name + extension)
        if os.path.exists(potential_path):
            return potential_path
    
    # If none found, raise exception
    raise FileNotFoundError(f"Image file not found: {img_name}\nTried paths include: {img_path} and various extension variants")


class UCEISMultiTaskDataset(Dataset):
    """UCEIS multi-task dataset class"""
    def __init__(self, csv_file, images_dir='./images', transform=None):
        self.data_frame = pd.read_csv(csv_file)
        self.images_dir = images_dir
        self.transform = transform
        
        # Check for NaN image paths
        nan_image_paths = self.data_frame[self.data_frame['image_path'].isna()]
        if len(nan_image_paths) > 0:
            print(f"Warning: Found {len(nan_image_paths)} records with NaN image_path in CSV file")
            print(f"   These records have indices: {list(nan_image_paths.index)}")
            print("   The program will terminate when trying to load these images")

    def __len__(self):
        return len(self.data_frame)

    def __getitem__(self, idx):
        if torch.is_tensor(idx):
            idx = idx.tolist()

        img_name = self.data_frame.iloc[idx, 2]  # image_path column
        patient_id = self.data_frame.iloc[idx, 1]  # patient_id column
        
        #print(f"Processing record {idx+1}: patient_id={patient_id}, image_path={img_name}")
        
        # Use helper function to find image file
        img_path = find_image_file(self.images_dir, img_name)
        
        # Read image
        image = Image.open(img_path).convert('RGB')
        
        # Get labels for three tasks
        vascular = self.data_frame.iloc[idx, 5]  # vascular column
        bleeding = self.data_frame.iloc[idx, 6]  # bleeding column
        ulceration = self.data_frame.iloc[idx, 7]  # ulceration column
        
        # Validate and limit label values to reasonable range
        vascular = max(0, min(2, int(vascular)))
        bleeding = max(0, min(3, int(bleeding)))
        ulceration = max(0, min(3, int(ulceration)))

        if self.transform:
            image = self.transform(image)

        return image, torch.tensor(vascular), torch.tensor(bleeding), torch.tensor(ulceration)


def evaluate_coatnet3_fold(fold=None, backbone_name='coatnet_3_rw_224', images_dir='./images', 
                          model_file=None, test_file=None):
    """Evaluate single fold CoAtNet-3 model"""
    print(f"===== CoAtNet-3 Model Evaluation =====")
    
    # File path handling
    if test_file:
        val_file = test_file
        print(f"Using specified test file: {val_file}")
    else:
        if fold is None:
            raise ValueError("Must provide either fold parameter or test_file parameter")
        val_file = f'kfold2/test_fold{fold}.csv'  # Use validation set for evaluation
        print(f"Using default test file: {val_file}")
    
    # Model file handling
    if model_file:
        if not os.path.exists(model_file):
            raise FileNotFoundError(f"Specified model file not found: {model_file}")
        print(f"Using specified model file: {model_file}")
    else:
        if fold is None:
            raise ValueError("Must provide either fold parameter or model_file parameter")
        # Try different model file naming conventions
        possible_model_files = [
            f'models/best_coatnet3_fold{fold}.pth',
            f'models/best_coatnet3_fold{fold}_no_simam.pth',
            f'models/best_coatnet3_fold{fold}_simam.pth'
        ]
        
        model_file = None
        for file_path in possible_model_files:
            if os.path.exists(file_path):
                model_file = file_path
                break
        
        if model_file is None:
            raise FileNotFoundError(f"Model file for fold {fold} not found. Tried paths: {possible_model_files}")
        print(f"Found model file: {model_file}")
    
    print(f"Using model file: {model_file}")
    
    # Check if data file exists
    if not os.path.exists(val_file):
        raise FileNotFoundError(f"Validation file not found: {val_file}")
    
    # Set device
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    print(f"Using device: {device}")
    
    # Data preprocessing - consistent with training
    val_transform = transforms.Compose([
        transforms.Resize((224, 224)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
    ])
    
    # Load validation dataset
    print("Loading validation dataset...")
    val_dataset = UCEISMultiTaskDataset(
        csv_file=val_file,
        images_dir=images_dir,
        transform=val_transform
    )
    
    # Create data loader
    val_loader = DataLoader(val_dataset, batch_size=32, shuffle=False, num_workers=0, drop_last=True)
    
    print(f"Validation set size: {len(val_dataset)}")
    print(f"Validation batch count: {len(val_loader)}")
    
    # Create CoAtNet-3 model
    print("Creating CoAtNet-3 model...")
    model = create_pure_coatnet_model(
        backbone_name=backbone_name,
        num_classes_per_task=(3, 4, 4),
        pretrained=False,  # No need for pretrained weights during evaluation
        dropout_rate=0.5
    )
    model = model.to(device)
    
    # Load trained model weights
    print("Loading model weights...")
    try:
        state_dict = torch.load(model_file, map_location=device)
        model.load_state_dict(state_dict)
        print("Model weights loaded successfully")
    except Exception as e:
        print(f"Failed to load model weights: {e}")
        return None
    
    model.eval()
    
    # Store prediction results and true labels
    all_vascular_preds = []
    all_vascular_targets = []
    all_bleeding_preds = []
    all_bleeding_targets = []
    all_ulceration_preds = []
    all_ulceration_targets = []
    all_image_names = []
    all_patient_ids = []  # Added: store patient_id
    
    # Store prediction probabilities (for further analysis)
    all_vascular_probs = []
    all_bleeding_probs = []
    all_ulceration_probs = []
    
    print("Starting evaluation...")
    
    with torch.no_grad():
        for batch_idx, (images, vascular_labels, bleeding_labels, ulceration_labels) in enumerate(val_loader):
            images = images.to(device)
            vascular_labels = vascular_labels.to(device)
            bleeding_labels = bleeding_labels.to(device)
            ulceration_labels = ulceration_labels.to(device)
            
            # Get current batch image names and patient_ids
            start_idx = batch_idx * val_loader.batch_size
            end_idx = min(start_idx + len(images), len(val_dataset))
            batch_image_names = [val_dataset.data_frame.iloc[i, 2] for i in range(start_idx, end_idx)]
            batch_patient_ids = [val_dataset.data_frame.iloc[i, 1] for i in range(start_idx, end_idx)]  # patient_id in column 1
            all_image_names.extend(batch_image_names)
            all_patient_ids.extend(batch_patient_ids)
            
            # Forward pass
            vascular_out, bleeding_out, ulceration_out = model(images)
            
            # Get prediction probabilities
            vascular_probs = torch.softmax(vascular_out, dim=1)
            bleeding_probs = torch.softmax(bleeding_out, dim=1)
            ulceration_probs = torch.softmax(ulceration_out, dim=1)
            
            # Get prediction results
            _, vascular_predicted = torch.max(vascular_out, 1)
            _, bleeding_predicted = torch.max(bleeding_out, 1)
            _, ulceration_predicted = torch.max(ulceration_out, 1)
            
            # Store results
            all_vascular_preds.extend(vascular_predicted.cpu().numpy())
            all_vascular_targets.extend(vascular_labels.cpu().numpy())
            all_bleeding_preds.extend(bleeding_predicted.cpu().numpy())
            all_bleeding_targets.extend(bleeding_labels.cpu().numpy())
            all_ulceration_preds.extend(ulceration_predicted.cpu().numpy())
            all_ulceration_targets.extend(ulceration_labels.cpu().numpy())
            
            # Store probabilities
            all_vascular_probs.extend(vascular_probs.cpu().numpy())
            all_bleeding_probs.extend(bleeding_probs.cpu().numpy())
            all_ulceration_probs.extend(ulceration_probs.cpu().numpy())
            
            if (batch_idx + 1) % 10 == 0:
                print(f"Processed {batch_idx + 1}/{len(val_loader)} batches")
    
    print("Evaluation completed, calculating metrics...")
    
    # Calculate metrics
    results = {}
    
    # Task label mapping
    task_labels = {
        'vascular': ['Normal', 'Mild', 'Severe'],
        'bleeding': ['None', 'Mild', 'Moderate', 'Severe'],
        'ulceration': ['None', 'Mild', 'Moderate', 'Severe']
    }
    
    # Vascular task evaluation
    vascular_acc = accuracy_score(all_vascular_targets, all_vascular_preds)
    vascular_precision, vascular_recall, vascular_f1, _ = precision_recall_fscore_support(
        all_vascular_targets, all_vascular_preds, average='weighted', zero_division=0)
    vascular_kappa = cohen_kappa_score(all_vascular_targets, all_vascular_preds)
    vascular_qwk = cohen_kappa_score(all_vascular_targets, all_vascular_preds, weights='quadratic')
    
    results['vascular'] = {
        'accuracy': vascular_acc,
        'precision': vascular_precision,
        'recall': vascular_recall,
        'f1_score': vascular_f1,
        'kappa': vascular_kappa,
        'qwk': vascular_qwk,
        'confusion_matrix': confusion_matrix(all_vascular_targets, all_vascular_preds),
        'classification_report': classification_report(
            all_vascular_targets, all_vascular_preds, 
            target_names=task_labels['vascular'], output_dict=True)
    }
    
    # Bleeding task evaluation
    bleeding_acc = accuracy_score(all_bleeding_targets, all_bleeding_preds)
    bleeding_precision, bleeding_recall, bleeding_f1, _ = precision_recall_fscore_support(
        all_bleeding_targets, all_bleeding_preds, average='weighted', zero_division=0)
    bleeding_kappa = cohen_kappa_score(all_bleeding_targets, all_bleeding_preds)
    bleeding_qwk = cohen_kappa_score(all_bleeding_targets, all_bleeding_preds, weights='quadratic')
    
    results['bleeding'] = {
        'accuracy': bleeding_acc,
        'precision': bleeding_precision,
        'recall': bleeding_recall,
        'f1_score': bleeding_f1,
        'kappa': bleeding_kappa,
        'qwk': bleeding_qwk,
        'confusion_matrix': confusion_matrix(all_bleeding_targets, all_bleeding_preds),
        'classification_report': classification_report(
            all_bleeding_targets, all_bleeding_preds, 
            target_names=task_labels['bleeding'], output_dict=True)
    }
    
    # Ulceration task evaluation
    ulceration_acc = accuracy_score(all_ulceration_targets, all_ulceration_preds)
    ulceration_precision, ulceration_recall, ulceration_f1, _ = precision_recall_fscore_support(
        all_ulceration_targets, all_ulceration_preds, average='weighted', zero_division=0)
    ulceration_kappa = cohen_kappa_score(all_ulceration_targets, all_ulceration_preds)
    ulceration_qwk = cohen_kappa_score(all_ulceration_targets, all_ulceration_preds, weights='quadratic')
    
    results['ulceration'] = {
        'accuracy': ulceration_acc,
        'precision': ulceration_precision,
        'recall': ulceration_recall,
        'f1_score': ulceration_f1,
        'kappa': ulceration_kappa,
        'qwk': ulceration_qwk,
        'confusion_matrix': confusion_matrix(all_ulceration_targets, all_ulceration_preds),
        'classification_report': classification_report(
            all_ulceration_targets, all_ulceration_preds, 
            target_names=task_labels['ulceration'], output_dict=True)
    }
    
    # Overall UCEIS accuracy evaluation (all three tasks predicted correctly)
    vascular_correct = np.array(all_vascular_preds) == np.array(all_vascular_targets)
    bleeding_correct = np.array(all_bleeding_preds) == np.array(all_bleeding_targets)
    ulceration_correct = np.array(all_ulceration_preds) == np.array(all_ulceration_targets)
    
    # Calculate overall accuracy
    overall_correct = vascular_correct & bleeding_correct & ulceration_correct
    overall_acc = np.mean(overall_correct.astype(float))
    
    # Calculate composite Kappa coefficient
    overall_targets = (np.array(all_vascular_targets) * 100 + 
                      np.array(all_bleeding_targets) * 10 + 
                      np.array(all_ulceration_targets))
    overall_preds = (np.array(all_vascular_preds) * 100 + 
                    np.array(all_bleeding_preds) * 10 + 
                    np.array(all_ulceration_preds))
    overall_kappa = cohen_kappa_score(overall_targets, overall_preds)
    overall_qwk = cohen_kappa_score(overall_targets, overall_preds, weights='quadratic')
    
    results['overall'] = {
        'accuracy': overall_acc,
        'kappa': overall_kappa,
        'qwk': overall_qwk
    }
    
    # Save detailed prediction results
    results_df = pd.DataFrame({
        'patient_id': all_patient_ids,  # Added: patient_id column
        'image_name': all_image_names,
        'vascular_actual': all_vascular_targets,
        'bleeding_actual': all_bleeding_targets,
        'ulceration_actual': all_ulceration_targets,
        'vascular_predicted': all_vascular_preds,
        'bleeding_predicted': all_bleeding_preds,
        'ulceration_predicted': all_ulceration_preds,
        'vascular_confidence': [max(probs) for probs in all_vascular_probs],
        'bleeding_confidence': [max(probs) for probs in all_bleeding_probs],
        'ulceration_confidence': [max(probs) for probs in all_ulceration_probs]
    })
    
    # Save results to CSV - using test filename with _result suffix
    os.makedirs('results', exist_ok=True)
    
    if fold is not None:
        # 5-fold cross-validation mode
        results_file = f'results/test_fold{fold}_result.csv'
    else:
        # Single model evaluation mode - extract test filename and add _result suffix
        test_filename = os.path.basename(val_file)
        test_name_without_ext = os.path.splitext(test_filename)[0]
        results_file = f'results/{test_name_without_ext}_result.csv'
    
    results_df.to_csv(results_file, index=False)
    print(f"Detailed prediction results saved to: {results_file}")
    
    # Print results
    print("=" * 80)
    if fold is not None:
        print(f"CoAtNet-3 Model Evaluation Results - Fold {fold+1}")
    else:
        print(f"Single Model CoAtNet-3 Evaluation Results")
    print("=" * 80)
    
    for task in ['vascular', 'bleeding', 'ulceration']:
        metrics = results[task]
        print(f"\n{task.capitalize()} Task:")
        print(f"  Accuracy: {metrics['accuracy']:.4f}")
        print(f"  Precision: {metrics['precision']:.4f}")
        print(f"  Recall: {metrics['recall']:.4f}")
        print(f"  Weighted F1 Score: {metrics['f1_score']:.4f}")
        print(f"  Kappa Coefficient: {metrics['kappa']:.4f}")
        print(f"  Quadratic Weighted Kappa (QWK): {metrics['qwk']:.4f}")
    
    # Print overall UCEIS results
    print(f"\nOverall UCEIS Task:")
    print(f"  Overall Accuracy: {results['overall']['accuracy']:.4f}")
    print(f"  Kappa Coefficient: {results['overall']['kappa']:.4f}")
    print(f"  Quadratic Weighted Kappa (QWK): {results['overall']['qwk']:.4f}")
    
    # Calculate average performance across tasks
    main_tasks = ['vascular', 'bleeding', 'ulceration']
    avg_accuracy = np.mean([results[task]['accuracy'] for task in main_tasks])
    avg_precision = np.mean([results[task]['precision'] for task in main_tasks])
    avg_recall = np.mean([results[task]['recall'] for task in main_tasks])
    avg_f1 = np.mean([results[task]['f1_score'] for task in main_tasks])
    avg_kappa = np.mean([results[task]['kappa'] for task in main_tasks])
    avg_qwk = np.mean([results[task]['qwk'] for task in main_tasks])
    
    print("\n" + "=" * 80)
    print("Average Performance Across Tasks:")
    print(f"  Average Accuracy: {avg_accuracy:.4f}")
    print(f"  Average Precision: {avg_precision:.4f}")
    print(f"  Average Recall: {avg_recall:.4f}")
    print(f"  Average Weighted F1 Score: {avg_f1:.4f}")
    print(f"  Average Kappa Coefficient: {avg_kappa:.4f}")
    print(f"  Average Quadratic Weighted Kappa (QWK): {avg_qwk:.4f}")
    print("=" * 80)
    
    return results, results_df


def plot_confusion_matrices(results, fold, save_dir='results'):
    """Plot confusion matrices"""
    os.makedirs(save_dir, exist_ok=True)
    
    task_labels = {
        'vascular': ['Normal', 'Mild', 'Severe'],
        'bleeding': ['None', 'Mild', 'Moderate', 'Severe'],
        'ulceration': ['None', 'Mild', 'Moderate', 'Severe']
    }
    
    fig, axes = plt.subplots(1, 3, figsize=(15, 4))
    
    for i, task in enumerate(['vascular', 'bleeding', 'ulceration']):
        cm = results[task]['confusion_matrix']
        labels = task_labels[task]
        
        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', 
                   xticklabels=labels, yticklabels=labels, ax=axes[i])
        axes[i].set_title(f'{task.capitalize()} Confusion Matrix')
        axes[i].set_xlabel('Predicted Label')
        axes[i].set_ylabel('True Label')
    
    plt.tight_layout()
    
    # Save confusion matrices, handle fold=None case
    if fold is not None:
        cm_file = f'{save_dir}/confusion_matrices_fold{fold}.png'
    else:
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        cm_file = f'{save_dir}/confusion_matrices_single_{timestamp}.png'
    
    plt.savefig(cm_file, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Confusion matrices saved to: {cm_file}")


def run_coatnet3_5fold_evaluation(backbone_name='coatnet_3_rw_224', images_dir='./images'):
    """Run CoAtNet-3 5-fold cross-validation evaluation"""
    print(" CoAtNet-3 5-Fold Cross-Validation Model Evaluation")
    print("=" * 80)
    
    # Check model files
    available_folds = []
    for fold in range(5):
        possible_files = [
            f'models/best_coatnet3_fold{fold}.pth',
            f'models/best_coatnet3_fold{fold}_no_simam.pth',
            f'models/best_coatnet3_fold{fold}_simam.pth'
        ]
        
        for file_path in possible_files:
            if os.path.exists(file_path):
                available_folds.append(fold)
                break
    
    if not available_folds:
        print(" No trained model files found")
        print("Please ensure run_improved_training.py has been executed to complete training")
        return
    
    print(f"Found {len(available_folds)} available model files: Fold(s) {[f+1 for f in available_folds]}")
    
    # Store results for all folds
    all_fold_results = {}
    successful_folds = []
    failed_folds = []
    
    start_time = datetime.datetime.now()
    
    for fold in available_folds:
        try:
            print(f"\n{'='*20} Evaluating Fold {fold+1} {'='*20}")
            
            fold_results, _ = evaluate_coatnet3_fold(fold, backbone_name, images_dir)
            
            if fold_results is not None:
                all_fold_results[fold] = fold_results
                successful_folds.append(fold)
                
                # Plot confusion matrices
                plot_confusion_matrices(fold_results, fold)
                
                print(f" Fold {fold+1} evaluation completed")
            else:
                failed_folds.append(fold)
                print(f" Fold {fold+1} evaluation failed")
                
        except Exception as e:
            print(f" Fold {fold+1} evaluation error: {e}")
            failed_folds.append(fold)
    
    end_time = datetime.datetime.now()
    total_duration = end_time - start_time
    
    # Only output completion information, no cross-fold average performance calculation
    if successful_folds:
        print(f"\n" + "=" * 80)
        print(f" CoAtNet-3 5-Fold Cross-Validation Evaluation Completed")
        print(f"=" * 80)
        
        print(f"Evaluation duration: {total_duration}")
        print(f"Successful folds: {len(successful_folds)}/5")
        print(f"Failed folds: {len(failed_folds)}/5")
        print(f"Successfully evaluated folds: {[fold+1 for fold in successful_folds]}")
        
        print(f"\nCoAtNet-3 5-Fold Cross-Validation Evaluation Completed!")
        print(f"   Detailed performance for each fold has been output separately during evaluation")
        
    else:
        print(f"\n All folds evaluation failed, please check model files and data")


def run_single_model_evaluation(model_file, test_file, result_excel=None, sheet_name=None, 
    backbone_name='coatnet_3_rw_224',
    images_dir='./images'):
    """Run single model evaluation"""
    print(f"CoAtNet-3 Single Model Evaluation")
    print(f"=" * 80)
    
    try:
        # Evaluate single model
        results, results_df = evaluate_coatnet3_fold(
            backbone_name=backbone_name,
            images_dir=images_dir,
            model_file=model_file,
            test_file=test_file
        )
        
        if results:
            print(f"Single model evaluation completed!")
            # Plot confusion matrices
            plot_confusion_matrices(results, 0, save_dir='results')
            
            # Write results to Excel file
            if result_excel and sheet_name:
                import os
                from openpyxl import load_workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Ensure directory exists
                excel_dir = os.path.dirname(result_excel)
                if excel_dir and not os.path.exists(excel_dir):
                    os.makedirs(excel_dir)
                
                try:
                    # Check if file exists
                    if os.path.exists(result_excel):
                        # Load existing Excel file
                        workbook = load_workbook(result_excel)
                        
                        # Delete existing worksheet if it exists
                        if sheet_name in workbook.sheetnames:
                            del workbook[sheet_name]
                            print(f"Existing worksheet deleted: {sheet_name}")
                    else:
                        # Create new Excel file
                        from openpyxl import Workbook
                        workbook = Workbook()
                        # Delete default Sheet if it exists
                        if 'Sheet' in workbook.sheetnames:
                            del workbook['Sheet']
                    
                    # Create new worksheet
                    worksheet = workbook.create_sheet(sheet_name)
                    
                    # Write dataframe to worksheet
                    for r_idx, row in enumerate(dataframe_to_rows(results_df, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            worksheet.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Save Excel file
                    workbook.save(result_excel)
                    print(f"Results successfully written to Excel file: {result_excel}")
                    print(f"   Worksheet name: {sheet_name}")
                except Exception as e:
                    print(f"Failed to write to Excel file: {e}")
                    import traceback
                    traceback.print_exc()
        else:
            print(f"Single model evaluation failed")
            
    except Exception as e:
        print(f"Single model evaluation error: {e}")
        import traceback
        traceback.print_exc()

# 2 3还可以
if __name__ == "__main__":
    fold = 0
    run_single_model_evaluation(
                model_file=fr'models\best_coatnet3_fold{fold}_no_simam.pth',
                test_file=r'kfold\test_ext.csv',
                result_excel = r'results\test_ext_result.xlsx',
                sheet_name = f"fold{fold}",
            )